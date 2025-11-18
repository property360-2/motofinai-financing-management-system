"""
Report generation utilities for PDF and Excel exports
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.apps import apps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class BaseReport:
    """Base class for report generation"""

    @staticmethod
    def generate_excel_response(filename):
        """Create HttpResponse for Excel file"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def style_header_row(worksheet, row=1):
        """Apply styling to header row"""
        for cell in worksheet[row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def auto_adjust_column_width(worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


class LoanReport(BaseReport):
    """Loan application reports"""

    @classmethod
    def generate_excel(cls, filters=None):
        """Generate Excel report of loan applications"""
        LoanApplication = apps.get_model('loans', 'LoanApplication')

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Loan Applications"

        # Headers
        headers = [
            'ID', 'Applicant Name', 'Email', 'Phone', 'Motor',
            'Loan Amount', 'Down Payment', 'Principal', 'Monthly Income',
            'Status', 'Created Date', 'Updated Date'
        ]
        ws.append(headers)
        cls.style_header_row(ws)

        # Get data
        queryset = LoanApplication.objects.select_related('motor').order_by('-submitted_at')
        if filters:
            queryset = queryset.filter(**filters)

        # Add data rows
        for loan in queryset:
            applicant_name = f"{loan.applicant_first_name} {loan.applicant_last_name}"
            ws.append([
                loan.id,
                applicant_name,
                loan.applicant_email,
                loan.applicant_phone,
                str(loan.motor) if loan.motor else 'N/A',
                float(loan.loan_amount or 0),
                float(loan.down_payment or 0),
                float(loan.principal_amount or 0),
                float(loan.monthly_income or 0),
                loan.get_status_display(),
                loan.submitted_at.strftime('%Y-%m-%d %H:%M'),
                loan.updated_at.strftime('%Y-%m-%d %H:%M'),
            ])

        # Auto-adjust columns
        cls.auto_adjust_column_width(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"loan_applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = cls.generate_excel_response(filename)
        response.write(output.getvalue())
        return response


class PaymentReport(BaseReport):
    """Payment reports"""

    @classmethod
    def generate_excel(cls, filters=None):
        """Generate Excel report of payments"""
        PaymentSchedule = apps.get_model('loans', 'PaymentSchedule')
        Payment = apps.get_model('payments', 'Payment')

        # Create workbook
        wb = Workbook()

        # Sheet 1: Payment Schedule
        ws1 = wb.active
        ws1.title = "Payment Schedule"

        headers1 = [
            'ID', 'Loan ID', 'Applicant', 'Sequence', 'Due Date',
            'Principal', 'Interest', 'Total Amount', 'Status'
        ]
        ws1.append(headers1)
        cls.style_header_row(ws1)

        queryset = PaymentSchedule.objects.select_related('loan_application').order_by('-due_date')
        if filters:
            queryset = queryset.filter(**filters)

        for schedule in queryset:
            ws1.append([
                schedule.id,
                schedule.loan_application.id,
                schedule.loan_application.applicant_first_name,
                schedule.sequence,
                schedule.due_date.strftime('%Y-%m-%d'),
                float(schedule.principal_amount),
                float(schedule.interest_amount),
                float(schedule.total_amount),
                schedule.get_status_display(),
            ])

        cls.auto_adjust_column_width(ws1)

        # Sheet 2: Payments Made
        ws2 = wb.create_sheet(title="Payments Made")

        headers2 = [
            'ID', 'Schedule ID', 'Applicant', 'Amount', 'Payment Date',
            'Reference', 'Recorded By', 'Recorded At'
        ]
        ws2.append(headers2)
        cls.style_header_row(ws2)

        payments = Payment.objects.select_related(
            'schedule__loan_application', 'recorded_by'
        ).order_by('-payment_date')

        for payment in payments:
            applicant_name = f"{payment.schedule.loan_application.applicant_first_name} {payment.schedule.loan_application.applicant_last_name}"
            ws2.append([
                payment.id,
                payment.schedule.id,
                applicant_name,
                float(payment.amount),
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.reference or 'N/A',
                payment.recorded_by.email if payment.recorded_by else 'System',
                payment.recorded_at.strftime('%Y-%m-%d %H:%M'),
            ])

        cls.auto_adjust_column_width(ws2)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"payments_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = cls.generate_excel_response(filename)
        response.write(output.getvalue())
        return response


class RiskReport(BaseReport):
    """Risk assessment reports"""

    @classmethod
    def generate_excel(cls):
        """Generate Excel report of risk assessments"""
        RiskAssessment = apps.get_model('risk', 'RiskAssessment')

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Risk Assessments"

        headers = [
            'ID', 'Loan ID', 'Applicant', 'Risk Score', 'Risk Level',
            'Credit Score', 'Missed Payments', 'Employment Status',
            'DTI Ratio', 'Created At'
        ]
        ws.append(headers)
        cls.style_header_row(ws)

        # Get data
        queryset = RiskAssessment.objects.select_related('loan_application').order_by('-score')

        for assessment in queryset:
            ws.append([
                assessment.id,
                assessment.loan_application.id,
                assessment.loan_application.applicant_first_name,
                float(assessment.score),
                assessment.get_risk_level_display(),
                assessment.credit_score or 'N/A',
                assessment.missed_payments,
                assessment.loan_application.get_employment_status_display(),
                float(assessment.debt_to_income_ratio or 0),
                assessment.calculated_at.strftime('%Y-%m-%d %H:%M'),
            ])

        cls.auto_adjust_column_width(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"risk_assessments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = cls.generate_excel_response(filename)
        response.write(output.getvalue())
        return response


class InventoryReport(BaseReport):
    """Inventory reports"""

    @classmethod
    def generate_excel(cls):
        """Generate Excel report of inventory"""
        Motor = apps.get_model('inventory', 'Motor')

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        headers = [
            'ID', 'Type', 'Brand', 'Model', 'Year', 'Chassis Number',
            'Color', 'Purchase Price', 'Status', 'Created At'
        ]
        ws.append(headers)
        cls.style_header_row(ws)

        # Get data
        queryset = Motor.objects.all().order_by('-created_at')

        for motor in queryset:
            ws.append([
                motor.id,
                motor.get_type_display(),
                motor.brand,
                motor.model_name,
                motor.year,
                motor.chassis_number,
                motor.color,
                float(motor.purchase_price),
                motor.status.title(),  # Use derived status property
                motor.created_at.strftime('%Y-%m-%d %H:%M'),  # Use created_at instead
            ])

        cls.auto_adjust_column_width(ws)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = cls.generate_excel_response(filename)
        response.write(output.getvalue())
        return response
